# ruff: noqa: E402
"""Backend tests for XLSX export generation."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys
from uuid import uuid4

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from data_curation.api import db, datasets
from data_curation.api.export_xlsx import build_dedoublonnage_xlsx, build_modification_xlsx
from data_curation.api.pg.curation_tx import update_entity_record
from data_curation.models import Intermarc
from .utils import _cluster_zone, _records_to_csv_bytes, create_zone


def _read_sheet_rows(xlsx_bytes: bytes):
    workbook = load_workbook(BytesIO(xlsx_bytes))
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append([value for value in row])
    return rows


def test_export_dedoublonnage_rows_match_cluster_flags():
    dataset_id = f"export-dedoublonnage-{uuid4().hex[:8]}"

    anchor_1 = "ark:/12148/cb1000000001"
    anchor_2 = "ark:/12148/cb1000000002"
    target_1 = "ark:/12148/cb2000000001"
    target_2 = "ark:/12148/cb2000000002"
    target_3 = "ark:/12148/cb2000000003"

    records = [
        {
            "id": "a1",
            "type": "Oeuvre",
            "intermarc": Intermarc(
                zones=[
                    create_zone("001", [("a", anchor_1, None)]),
                    _cluster_zone(target_1, note="Clusterisation manuelle", affected="manual"),
                ]
            ).to_json_string(),
        },
        {
            "id": "a2",
            "type": "Oeuvre",
            "intermarc": Intermarc(
                zones=[
                    create_zone("001", [("a", anchor_2, None)]),
                    _cluster_zone(target_2, note="Clusterisation script", affected="script"),
                    _cluster_zone(target_3, note="Clusterisation script", affected="script"),
                ]
            ).to_json_string(),
        },
        {
            "id": "a3",
            "type": "Oeuvre",
            "intermarc": Intermarc(
                zones=[
                    create_zone("001", [("a", "ark:/12148/cb1000000003", None)]),
                    _cluster_zone("ark:/12148/cb2000000004", note="Autre note", affected="manual"),
                    _cluster_zone("ark:/12148/cb2000000005", note="Clusterisation manuelle", affected=""),
                ]
            ).to_json_string(),
        },
    ]

    datasets.ensure_dataset(dataset_id, title="export dedoublonnage")
    db.ingest_csv(_records_to_csv_bytes(records), dataset_id)

    payload = build_dedoublonnage_xlsx(dataset_id)
    rows = _read_sheet_rows(payload)

    assert rows == [
        ["RemplacementSimple", anchor_1, target_1],
        ["RemplacementMultiple", anchor_2, f"{target_2},{target_3}"],
    ]


def test_export_modifications_computes_add_delete_edit():
    dataset_id = f"export-modifications-{uuid4().hex[:8]}"

    ark = "ark:/12148/cb3000000000"
    original = Intermarc(
        zones=[
            create_zone("001", [("a", ark, None)]),
            create_zone("150", [("a", "Original Title", None)]),
            create_zone("245", [("b", "Old Subtitle", None)]),
            create_zone("300", [("a", "Base", None)]),
            _cluster_zone("ark:/12148/cb9000000000", note="Clusterisation manuelle", affected="manual"),
        ]
    )

    datasets.ensure_dataset(dataset_id, title="export modifications")
    db.ingest_csv(
        _records_to_csv_bytes([{"id": "r1", "type": "Oeuvre", "intermarc": original.to_json_string()}]),
        dataset_id,
    )

    updated = Intermarc(
        zones=[
            create_zone("001", [("a", ark, None)]),
            create_zone("150", [("a", "Updated Title", None)]),
            create_zone("300", [("a", "Base", None)]),
            create_zone("500", [("a", "New note", None)]),
            _cluster_zone("ark:/12148/cb9000000000", note="Clusterisation manuelle", affected="manual"),
        ]
    )
    update_entity_record(dataset_id, record_id="r1", type_raw="Oeuvre", intermarc=updated)

    payload = build_modification_xlsx(dataset_id)
    rows = _read_sheet_rows(payload)

    assert rows == [
        [ark, "ModifZone", "150", "$a Original Title", "$a Updated Title", None],
        [ark, "SupprZone", "245", "$b Old Subtitle", None, None],
        [ark, "AjoutZone", "500", None, "$a New note", None],
    ]
