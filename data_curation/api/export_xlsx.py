"""Generate XLSX exports for dataset curation outputs."""

from __future__ import annotations

import json
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from data_curation.api.pg.session import db_session, statement_timeout
from data_curation.models import Intermarc, Zone


CLUSTER_NOTES = {"Clusterisation manuelle", "Clusterisation script"}
CLUSTER_FLAGS = {"manual", "script"}

ROOT_DIR = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = ROOT_DIR / "documentation" / "export_templates"
DEDUP_TEMPLATE = TEMPLATE_DIR / "templateDedoublonnage.xlsx"
MOD_TEMPLATE = TEMPLATE_DIR / "templateModification.xlsx"


@dataclass(frozen=True)
class _ZoneSnapshot:
    code: str
    subfields: Tuple[Tuple[str, str], ...]
    display: str
    index: int


def _load_intermarc(payload: object | None) -> Intermarc:
    if payload is None:
        return Intermarc(zones=[])
    if isinstance(payload, str):
        return Intermarc.from_json_string(payload)
    return Intermarc.from_json_string(json.dumps(payload, ensure_ascii=False))


def _is_curated_cluster_zone(zone: Zone) -> bool:
    if zone.code != "90F":
        return False
    flag = (zone.affected_by_curation or "").strip().lower()
    if flag not in CLUSTER_FLAGS:
        return False
    note = next((sub.valeur for sub in zone.sousZones if sub.code == "90F$q"), None)
    if not note or note.strip() not in CLUSTER_NOTES:
        return False
    return True


def _extract_cluster_targets(zone: Zone) -> List[str]:
    targets: List[str] = []
    for sub in zone.sousZones:
        if sub.code == "90F$3" and str(sub.valeur).strip():
            targets.append(str(sub.valeur).strip())
    return targets


def _filter_curated_cluster_zones(zones: Sequence[Zone]) -> List[Zone]:
    return [zone for zone in zones if not _is_curated_cluster_zone(zone)]


def _extract_subfields(zone: Zone) -> List[Tuple[str, str]]:
    if zone.sousZones:
        return [(sub.code, sub.valeur) for sub in zone.sousZones]
    if zone.field_compact_value:
        try:
            parsed = json.loads(zone.field_compact_value)
            entries = parsed.get("sousZones", []) if isinstance(parsed, dict) else []
            if isinstance(entries, list) and entries:
                subfields: List[Tuple[str, str]] = []
                for entry in entries:
                    if not isinstance(entry, dict):
                        continue
                    code = str(entry.get("code", "") or "")
                    value = str(entry.get("valeur", "") or "")
                    subfields.append((code, value))
                if subfields:
                    return subfields
        except json.JSONDecodeError:
            pass
        return [("__compact__", str(zone.field_compact_value))]
    return []


def _format_zone_value(subfields: Sequence[Tuple[str, str]]) -> str:
    if not subfields:
        return ""
    if len(subfields) == 1 and subfields[0][0] == "__compact__":
        return subfields[0][1]
    parts: List[str] = []
    for code, value in subfields:
        label = code.split("$", 1)[1] if "$" in code else code
        label = label or code
        if value:
            parts.append(f"${label} {value}")
        else:
            parts.append(f"${label}")
    return " ".join(parts)


def _zone_snapshot(zone: Zone, index: int) -> _ZoneSnapshot:
    subfields = tuple(_extract_subfields(zone))
    return _ZoneSnapshot(
        code=zone.code,
        subfields=subfields,
        display=_format_zone_value(subfields),
        index=index,
    )


def _subfield_labels(subfields: Sequence[Tuple[str, str]]) -> List[str]:
    labels: List[str] = []
    for code, _ in subfields:
        if code == "__compact__":
            labels.append("__compact__")
        elif "$" in code:
            labels.append(code.split("$", 1)[1])
        else:
            labels.append(code)
    return labels


def _similarity(left: _ZoneSnapshot, right: _ZoneSnapshot) -> float:
    if not left.subfields and not right.subfields:
        if left.display and right.display:
            return 1.0 if left.display == right.display else 0.0
        return 0.0
    if not left.subfields or not right.subfields:
        return 0.0
    labels_left = _subfield_labels(left.subfields)
    labels_right = _subfield_labels(right.subfields)
    overlap = sum((Counter(labels_left) & Counter(labels_right)).values())
    if overlap == 0:
        return 0.0
    return (2.0 * overlap) / (len(labels_left) + len(labels_right))


def _value_overlap(left: _ZoneSnapshot, right: _ZoneSnapshot) -> int:
    left_pairs = Counter(left.subfields)
    right_pairs = Counter(right.subfields)
    return sum((left_pairs & right_pairs).values())


def _match_zone_pairs(
    originals: Sequence[_ZoneSnapshot],
    updated: Sequence[_ZoneSnapshot],
    *,
    threshold: float = 0.5,
) -> Tuple[List[Tuple[_ZoneSnapshot, _ZoneSnapshot]], List[_ZoneSnapshot], List[_ZoneSnapshot]]:
    pairs: List[Tuple[float, int, int, int]] = []
    for i, left in enumerate(originals):
        for j, right in enumerate(updated):
            score = _similarity(left, right)
            if score < threshold:
                continue
            overlap = _value_overlap(left, right)
            pairs.append((score, overlap, i, j))
    pairs.sort(key=lambda item: (-item[0], -item[1], item[2], item[3]))

    matched_left: set[int] = set()
    matched_right: set[int] = set()
    matches: List[Tuple[_ZoneSnapshot, _ZoneSnapshot]] = []
    for _, _, i, j in pairs:
        if i in matched_left or j in matched_right:
            continue
        matched_left.add(i)
        matched_right.add(j)
        matches.append((originals[i], updated[j]))

    remaining_left = [snap for idx, snap in enumerate(originals) if idx not in matched_left]
    remaining_right = [snap for idx, snap in enumerate(updated) if idx not in matched_right]
    return matches, remaining_left, remaining_right


def _prepare_template(path: Path):
    workbook = load_workbook(path)
    sheet = workbook.active
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    return workbook, sheet


def build_dedoublonnage_xlsx(dataset_id: str) -> bytes:
    """Build the deduplication XLSX export for a dataset."""
    rows: List[List[Optional[str]]] = []
    with db_session() as conn, statement_timeout(conn, 120_000):
        dataset_rows = conn.execute(
            "SELECT record_id, ark, record FROM entity WHERE dataset_id=%s ORDER BY record_id",
            (dataset_id,),
        ).fetchall()

    for row in dataset_rows:
        intermarc = _load_intermarc(row.get("record"))
        targets: List[str] = []
        for zone in intermarc.zones:
            if not _is_curated_cluster_zone(zone):
                continue
            for target in _extract_cluster_targets(zone):
                if target not in targets:
                    targets.append(target)
        if not targets:
            continue
        anchor_id = row.get("ark") or row.get("record_id") or ""
        if len(targets) == 1:
            rows.append(["RemplacementSimple", str(anchor_id), targets[0]])
        else:
            rows.append(["RemplacementMultiple", str(anchor_id), ",".join(targets)])

    workbook, sheet = _prepare_template(DEDUP_TEMPLATE)
    for line in rows:
        sheet.append(line)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_modification_xlsx(dataset_id: str) -> bytes:
    """Build the modification XLSX export comparing original vs current Intermarc records."""
    modifications: List[List[Optional[str]]] = []
    with db_session() as conn, statement_timeout(conn, 120_000):
        dataset_rows = conn.execute(
            "SELECT record_id, ark, record, original_record FROM entity WHERE dataset_id=%s ORDER BY record_id",
            (dataset_id,),
        ).fetchall()

    for row in dataset_rows:
        original_payload = row.get("original_record") or row.get("record")
        current_payload = row.get("record")
        original = _load_intermarc(original_payload)
        current = _load_intermarc(current_payload)

        original_zones = _filter_curated_cluster_zones(original.zones)
        current_zones = _filter_curated_cluster_zones(current.zones)

        original_snaps = [_zone_snapshot(zone, idx) for idx, zone in enumerate(original_zones)]
        current_snaps = [_zone_snapshot(zone, idx) for idx, zone in enumerate(current_zones)]

        tags = sorted({snap.code for snap in original_snaps} | {snap.code for snap in current_snaps})
        record_id = row.get("ark") or row.get("record_id") or ""

        for tag in tags:
            old_list = [snap for snap in original_snaps if snap.code == tag]
            new_list = [snap for snap in current_snaps if snap.code == tag]
            matches, remaining_old, remaining_new = _match_zone_pairs(old_list, new_list)

            for old_snap, new_snap in matches:
                if old_snap.display != new_snap.display:
                    modifications.append(
                        [
                            str(record_id),
                            "ModifZone",
                            tag,
                            old_snap.display or None,
                            new_snap.display or None,
                            None,
                        ]
                    )

            for old_snap in remaining_old:
                modifications.append(
                    [
                        str(record_id),
                        "SupprZone",
                        tag,
                        old_snap.display or None,
                        None,
                        None,
                    ]
                )

            for new_snap in remaining_new:
                modifications.append(
                    [
                        str(record_id),
                        "AjoutZone",
                        tag,
                        None,
                        new_snap.display or None,
                        None,
                    ]
                )

    workbook, sheet = _prepare_template(MOD_TEMPLATE)
    for line in modifications:
        sheet.append(line)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
